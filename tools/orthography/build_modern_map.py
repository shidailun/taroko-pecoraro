"""Build the word-level modern-spelling map for the Pecoraro dictionary app.

Pipeline (per distinct displayed Truku token):
  1. identity check    — original spelling already attested in the modern Truku
                         omnibus -> keep as-is (tier "id")
  2. candidate gen     — per-occurrence x>h / o>u / l>r branching, plus contextual
                         transforms (final -e>-i, -ae/-ai>-ay, -ao>-aw, -ea/-ia>-iya,
                         d>j before i, intervocalic u>w, e>i) applied on top
  3. attestation filter— candidate must exist in omnibus Words (strong) or as a
                         token of an omnibus Sentence (weak)
  4. gloss check       — Chinese gloss of the Pecoraro form vs the omnibus word's
                         Chinese gloss; a real overlap promotes to tier "A"
  5. tiering           — A  = Words-attested + gloss-confirmed
                         B  = unique attested candidate (no gloss available/needed)
                         C  = multiple attested candidates, gloss can't decide ->
                              NOT auto-applied; written to review file
Output:
  tools/orthography/modern_map.json   full evidence, all tiers incl. review
  site/modern_map.js                  window.MODERN_MAP, tiers id/A/B only
Superseded: site/keep_words.js (identity entries now live in the map).
"""
import json, re, unicodedata, collections, itertools, os, difflib

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.dirname(os.path.dirname(HERE))
ENTRIES = os.path.join(BASE, "site", "entries.js")
OMNIDIR = r"C:/Users/user/OneDrive - Lingnan University/Desktop/SeediqPro/dictionaries/omnibus"
OMNIBUS = OMNIDIR + "/Truku_Omnibus.xlsx"
TGDAYA = OMNIDIR + "/Omnibus.xlsx"          # Tgdaya omnibus (Words: word col 1, 華語 col 5)
TODA = OMNIDIR + "/toda.xlsx"               # no header row: word col 0, 中文 col 1

OVERRIDE_KEYS = {
    "klui", "mklui", "nklui", "tklui", "sklui", "msklui", "psklui",
    "mnsklui", "snklui", "mnklui", "kui", "mskui", "kskui", "ktui", "kmtui",
    "mktui", "bkui", "bukui", "mukui", "mkui", "mkbukui", "bklui", "bq'lui",
    "tutui", "mtutui", "dui", "dmui", "mdui", "mddui", "pdui", "sdui",
    "mndui", "mpdui", "xbui", "xmbui", "pxbui", "xnbui", "m'xapui", "mapui",
}

# Pecoraro types two elision marks, ' and ", and both sit INSIDE a word: page 47
# has BL'NGA and B"LO four lines apart, and Tmb"lo / knta"to / pn"lu keep the
# double mark right through a paradigm. A tokenizer that breaks on " turns one
# word into two fragments and maps each fragment separately, so " is a word
# character here and folds to ' — the same rule wordKey() applies in app.js.
TOKEN_RE = re.compile(r"[A-Za-zÀ-ÿłŁʔ'’ʼ\"]+")
ZH_RE = re.compile(r"[^一-鿿]")

def tkey(t):
    return (t.lower().replace("’", "'").replace("ʼ", "'").replace('"', "'")
            .replace("ʔ", "'").replace("ł", "l"))

ELISION_RE = re.compile("['’ʼ\"ʔ]")

def norm(w):
    # Pecoraro's ç is modern x (tunuç>tunux, otoç>utux) — map before NFD strips it
    w = (w or "").replace("ç", "x").replace("Ç", "X")
    w = unicodedata.normalize("NFD", w)
    w = "".join(c for c in w if unicodedata.category(c) != "Mn")
    # ł is his barred l (Małi vs MAI) and ʔ a glottal stop; without the fold the
    # final a-z filter would delete them and norm małi to "mai".
    w = w.lower().replace("ł", "l").replace("ʔ", "")
    w = w.replace("'", "").replace("’", "").replace("ʼ", "").replace("-", "")
    return re.sub(r"[^a-z]", "", w)

def plain(t):
    """His token with the diacritics dropped. ç is modern x (tunuç>tunux) so it is
    converted rather than stripped; every other mark he uses writes a vowel quality
    modern Truku does not spell (lamil/lämil, isu/isò, jiyan/diyán)."""
    t = t.replace("ç", "x").replace("Ç", "X").replace("ł", "l").replace("Ł", "L")
    t = unicodedata.normalize("NFD", t)
    return "".join(c for c in t if unicodedata.category(c) != "Mn")

def zh_clean(s):
    return ZH_RE.sub("", s or "")

# ---------------- corpus ----------------
def load_corpus():
    src = open(ENTRIES, encoding="utf-8").read()
    src = src[src.index("window.ENTRIES =") + len("window.ENTRIES ="):].strip().rstrip(";")
    entries = json.loads(src)
    tokens = collections.Counter()       # lowercase raw token -> display frequency
    glosses = collections.defaultdict(set)   # norm(form) -> zh gloss strings
    families = []                        # per entry: list of member tokens (hw+subs)
    # Capitalization as he wrote it, kept per token key. A word that NEVER appears
    # lowercase anywhere in 5,438 example sentences is a personal name, which is
    # what the collective d- in tier D needs to know.
    cased = collections.defaultdict(collections.Counter)
    def take(text):
        for t in TOKEN_RE.findall(text or ""):
            tokens[tkey(t)] += 1
            cased[tkey(t)][t] += 1
    # His root tags hold Truku too, and the census was blind to them: 443 of the
    # 1,850 tags contain the root mark and go through linkifyTruku on screen, where
    # they hold 103 token types no other field has — nearly all his own bracketed
    # variant spelling of the headword, `(LOKUS) (R)` under LUKUS, `(SYEQA - SYAQA)`
    # under SIQA, `(BQXOS)` under BQLOS. Not being in the census, no tier could
    # reach them and every one fell to the char rules: TNQDO's `(= R. ? - R. =
    # L'QDO ?)` printed a green RQDU next to the brown RQDUG of the entry it points
    # at. The gate below is tagHtml()'s, mirrored: a tag with no root mark renders
    # as plain grey text and is NOT Truku, the mark itself and the French `de`/`du`
    # that introduces a proposed root are stripped, and TAG_PROSE — his thirty-odd
    # tags that are a French remark rather than a form ("QALAO est plus probable")
    # — is filtered out. Mirrored, so it must be kept in step with app.js.
    # These join `tokens` only, never a family: a bracketed variant is a spelling he
    # is unsure of, not an inflection, so it earns attestation and elision-twin
    # evidence but must not seed a projection.
    #
    # And they are enrolled ONLY as new types — a token the census already has keeps
    # the count it had. The passes turn out to be order-dependent on frequency (they
    # walk `tokens` most-common-first, and each tier reads what earlier ones already
    # resolved), so merely bumping a count re-orders the walk and can flip a decision
    # somewhere else entirely. Bumping them cost two correct tier-E readings: GAGWI's
    # `gmagwi` went from gmeeguy to *ggmeeguy — analysed as g+magwi when it is g-m-agwi
    # infixed — and G'LEQ's `pgleqe` from pgriqi to *pgliqi, on a root the modern
    # language writes with r (mgriq 轉動). Neither token is in a tag. A marginal
    # source must not be able to reach across the book like that.
    TAG_PROSE = set((
        "de du des la le les ce cette et en un une avec sans dans sous tous plus "
        "peu tres doute pluriel travers tordu inconnue probable probablement "
        "parfois relation scie souvent reduit a fermer suie passer qui passe "
        "uniquement connu forme suivante derive derives precedent contraction "
        "faire venir manger variante etant escamote places terme semble vraie "
        "chinois chinoise est nb sy prefixe suffixe").split())
    ROOT_MARK = re.compile(r"(^|[\s(=-])R\.?($|[\s)?=.-])")
    ROOT_MARK_G = re.compile(r"(^|[\s(=-])R\.?(?=$|[\s)?=.-])")
    TAG_FRENCH = re.compile(r"(^|[\s(=-])d[eu](?=\s)")
    tag_seen = collections.Counter()
    def take_tag(tag):
        if not tag or tag in ("(R)", "(R.)") or not ROOT_MARK.search(tag):
            return
        rest = TAG_FRENCH.sub(r"\1", ROOT_MARK_G.sub(
            r"\1", re.sub(r"\(\s*R\.?\s*\)", " ", tag)))
        for t in TOKEN_RE.findall(rest):
            k = tkey(t)
            if len(k) < 2 or plain(k).strip("'") in TAG_PROSE:
                continue
            tag_seen[k] += 1
            if k not in tag_cased:
                tag_cased[k] = t
    tag_cased = {}
    # A capital in the middle of a sentence is the one place Pecoraro's own
    # typography names a proper noun for us. Combined with "never seen lowercase
    # anywhere", it is what separates Sibal the man from sibal the word — and the
    # l>r rule must not touch him: he came out of the app as Sibar.
    midcap = collections.Counter()
    def take_ex(text):
        take(text)
        if not text:
            return
        for m in TOKEN_RE.finditer(text):
            w = m.group(0)
            if not w[:1].isupper():
                continue
            pre = text[:m.start()].rstrip()
            if not pre or pre[-1] in ".!?…§:":
                continue
            midcap[tkey(w)] += 1
    # The words of an entry's own example sentences, kept beside its family but
    # NOT in it: a sentence is mostly other people's words, so they must never be
    # treated as inflections of this root. Tier E tests them for containment of a
    # stem this entry has already resolved, which is a far narrower claim.
    ex_fams = []
    heads = []
    # The words of the entries Pecoraro tags [emprunt jap./chin.]. He romanizes
    # these faithfully to the source (Japanese o stays o: SATO, DOKU, OTOBAI),
    # which is a different system from the one he uses for Truku, where his o is
    # the sound modern orthography writes u. They therefore need their own pass —
    # and, more urgently, they must be kept away from the attestation tiers. The
    # modern wordlists dropped nearly the whole loan stratum in favour of native
    # coinages (lumak for tobacco, mtgsa for teacher, tluan for table), so a loan
    # that "matches" a modern word is matching a homonym: TOKE 時計 was confirmed
    # as tuki because tuki occurs 312 times in speech — meaning 抵銷.
    #
    # Several loan entries are compounds with a native word in them — Sapax kensat
    # "police station", Tama denki "electric pole", BALA-NO-XANA "bara no hana" —
    # and taking the headword apart naively enrolled sapax (375 occurrences in the
    # book), tama (131) and the Japanese genitive no, which is spelled exactly like
    # the Truku particle no (210). Those are not loans, and a class pass that
    # outranks attestation must not be allowed to decide them. So a multi-token
    # source contributes only the tokens that occur nowhere outside the loan
    # entries; a single-token one is the loan itself and always counts.
    loan_srcs, native_tokens = [], set()
    # Our own digitization tags a man's or woman's name outright — `name (m)` 137
    # times, `name (f)` 87 — and nothing here ever read it. Tier N reconstructs the
    # same fact from capitalisation statistics, so it reaches only the names he
    # happened to put in a sentence, and l>r renamed the rest: LAKAX came out
    # Rakah, SOBIL Subir, TOLI Turi. The tag IS the evidence; take it.
    #
    # Two restrictions, or the seed does damage. A name whose token is also some
    # other entry's headword or sub-form is excluded: Truku names ARE nouns
    # (LONGAI 猴子, XALONG 松樹, PALAS, KALAO), the noun is the entry carrying the
    # gloss, and one bare token cannot render two ways. And `name (.., jp)` is
    # excluded because his Japanese romanization is a different system — the tier-J
    # comment above says his Japanese o stays o (SATO, DOKU) — so whether Toro is
    # turu or toro is a question about Japanese, not about his Truku spelling.
    name_first, nonname_forms = set(), set()
    for e in entries:
        tg = e.get("tag") or ""
        isname = re.search(r"\bname\b", tg, re.I) and "jp" not in tg.lower()
        forms = [e.get("hw")] + [s.get("form") for s in e.get("subs", [])]
        if isname:
            hw0 = TOKEN_RE.findall(re.split(r"[(\[=]", e.get("hw") or "")[0])
            if hw0:
                name_first.add(tkey(hw0[0]))
        if not re.search(r"\bname\b", tg, re.I):
            for f in forms:
                nonname_forms.update(tkey(t) for t in
                                     TOKEN_RE.findall(re.split(r"[(\[=]", f or "")[0]))
    name_heads = name_first - nonname_forms
    for e in entries:
        hwt = TOKEN_RE.findall(e.get("hw") or "")
        if hwt:
            heads.append(tkey(hwt[0]))
        srcs = [e.get("hw"), e.get("paradigm")] + \
               [s.get(k) for s in e.get("subs", []) for k in ("form", "paradigm")]
        # three tag spellings, all his: [emprunt jap./chin.] x121, (J) on KENSAT,
        # (J.?) on BAKET — the query mark is his own doubt about the etymology,
        # not a different category.
        if re.search(r"emprunt|^\(J", e.get("tag") or ""):
            loan_srcs += srcs
        else:
            for src in srcs + [x.get("t") for x in e.get("examples", [])] + \
                       [x.get("t") for s in e.get("subs", []) for x in s.get("examples", [])]:
                native_tokens.update(tkey(t) for t in TOKEN_RE.findall(src or ""))
        take(e.get("hw")); take(e.get("crossRef")); take(e.get("paradigm"))
        take_tag(e.get("tag"))
        for x in e.get("examples", []): take_ex(x.get("t"))
        hz = zh_clean(e.get("zh", ""))
        if e.get("hw") and hz: glosses[norm(e["hw"])].add(hz)
        fam = [tkey(t) for t in TOKEN_RE.findall(e.get("hw") or "")]
        ext = [tkey(t) for x in e.get("examples", [])
               for t in TOKEN_RE.findall(x.get("t") or "")]
        # paradigm lines (° gmalax, malax...) are inflections of THIS root — family
        for t in TOKEN_RE.findall(e.get("paradigm") or ""):
            fam.append(tkey(t))
        for s in e.get("subs", []):
            take(s.get("form")); take(s.get("paradigm"))
            for x in s.get("examples", []): take_ex(x.get("t"))
            sz = zh_clean(s.get("zh", ""))
            if s.get("form") and sz: glosses[norm(s["form"])].add(sz)
            for t in TOKEN_RE.findall(s.get("form") or ""):
                fam.append(tkey(t))
            for t in TOKEN_RE.findall(s.get("paradigm") or ""):
                fam.append(tkey(t))
            ext += [tkey(t) for x in s.get("examples", [])
                    for t in TOKEN_RE.findall(x.get("t") or "")]
        if len(fam) > 1:
            families.append(fam)
        if fam and ext:
            ex_fams.append((fam, ext))
    # new types only — see take_tag
    for k, n in tag_seen.items():
        if k not in tokens:
            tokens[k] += n
            cased[k][tag_cased[k]] += n
    loan_tokens = set()
    for src in loan_srcs:
        ts = [tkey(t) for t in TOKEN_RE.findall(src or "")]
        loan_tokens.update(ts if len(ts) == 1 else [t for t in ts if t not in native_tokens])
    return (tokens, glosses, families, cased, ex_fams, midcap, heads, loan_tokens,
            name_heads)

# ---------------- omnibus ----------------
def load_omnibus():
    import openpyxl
    wb = openpyxl.load_workbook(OMNIBUS, read_only=True, data_only=True)
    word_gloss = collections.defaultdict(set)    # norm -> zh glosses
    word_raw = {}                                # norm -> canonical modern spelling
    for r in wb["Words"].iter_rows(min_row=2, values_only=True):
        w, zh = r[1], r[2]
        if not w: continue
        toks = [t for t in re.split(r"[^A-Za-z']+", str(w)) if t]
        for i, t in enumerate(toks):
            n = norm(t)
            if not n: continue
            if n not in word_raw or (i == 0 and len(toks) == 1):
                word_raw[n] = t.lower()
            if zh:
                word_gloss[n].add(zh_clean(str(zh)))
    sent_raw = collections.Counter()             # (norm, raw) counts
    for r in wb["Sentences"].iter_rows(min_row=2, values_only=True):
        if not r[1]: continue
        for t in re.split(r"[^A-Za-z']+", str(r[1])):
            n = norm(t)
            if len(n) >= 2:
                sent_raw[(n, t.lower())] += 1
    sent_best = {}
    for (n, raw), c in sorted(sent_raw.items(), key=lambda kv: -kv[1]):
        sent_best.setdefault(n, raw)
    return word_raw, word_gloss, sent_best

# ---------------- spoken corpus ----------------
SPOKEN = r"C:/dev/ILRDF/ILRDF_texts.xlsx"
SPOKEN_CACHE = os.path.join(HERE, "spoken_truku.json")

def load_spoken():
    """Running Truku, not a wordlist: 47,517 transcribed utterances / 277,014
    tokens across the ILRDF/klokah/ithuan collections. The omnibus is a
    dictionary, so it is missing exactly the words a dictionary tends to skip —
    personal names (Sibal, Wilang, Iwal), particles, and the shapes an inflected
    root actually takes in a sentence. Frequency is kept because a hapax in an
    ASR transcript is as likely to be a mis-hearing as a word."""
    if os.path.exists(SPOKEN_CACHE) and os.path.getmtime(SPOKEN_CACHE) > os.path.getmtime(SPOKEN):
        with open(SPOKEN_CACHE, encoding="utf-8") as f:
            return collections.Counter(json.load(f))
    import openpyxl
    wb = openpyxl.load_workbook(SPOKEN, read_only=True, data_only=True)
    freq = collections.Counter()
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        hdr = next(it, None)
        if not hdr or "dialect" not in hdr or "transcript" not in hdr:
            continue
        di, ti = hdr.index("dialect"), hdr.index("transcript")
        for r in it:
            if not r[di] or str(r[di]).strip() != "Truku" or not r[ti]:
                continue
            for w in re.split(r"[^A-Za-z']+", str(r[ti]).lower()):
                n = norm(w)
                if len(n) >= 2:
                    freq[n] += 1
    with open(SPOKEN_CACHE, "w", encoding="utf-8", newline="\n") as f:
        json.dump(dict(freq), f, ensure_ascii=False)
    return freq

# ---------------- sister dialects ----------------
def load_sisters():
    """Toda + Tgdaya lexicons, norm -> zh glosses. Used ONLY to validate which
    generated candidate is the right shape — never as a source of spellings.
    Toda orthography is Truku-like: index as-is plus an o>u fold. Tgdaya
    differs by regular correspondences (l>r, o>u, seediq>seejiq d>j / t>c
    before i), so each Tgdaya word indexes under every combination of those
    folds; when a Pecoraro token yields e.g. both an l- and an r-candidate with
    the same sister gloss, they tie and the token stays unmapped (review)
    rather than guessing."""
    import openpyxl

    def tg_folds(n):
        outs = {n}
        for f in (lambda w: w.replace("l", "r"),
                  lambda w: w.replace("o", "u"),
                  lambda w: re.sub(r"d(?=i)", "j", w),
                  lambda w: re.sub(r"t(?=i)", "c", w)):
            outs |= {f(w) for w in outs}
        return outs

    def strips(v):
        """Also index affix-stripped cores (>=5 chars) so a differently-derived
        cognate (Pecoraro baxang ~ Tgdaya qbahang/mbahang) can still validate
        the shared stem. Gloss agreement remains the gatekeeper."""
        outs = {v}
        for k in (1, 2, 3):
            if len(v) - k >= 5:
                outs.add(v[k:])
        return outs

    toda_g = collections.defaultdict(set)
    tg_g = collections.defaultdict(set)
    wb = openpyxl.load_workbook(TODA, read_only=True, data_only=True)
    for r in wb.worksheets[0].iter_rows(min_row=1, values_only=True):
        if not r or not r[0]: continue
        z = zh_clean(str(r[1] or ""))
        if not z: continue
        for tk in re.split(r"[^A-Za-z']+", str(r[0])):
            n = norm(tk)
            if len(n) >= 2:
                for base in (n, n.replace("o", "u")):
                    for v in strips(base):
                        toda_g[v].add(z)
    wb = openpyxl.load_workbook(TGDAYA, read_only=True, data_only=True)
    for r in wb["Words"].iter_rows(min_row=2, values_only=True):
        w, z = r[1], r[5]
        if not w: continue
        z = zh_clean(str(z or ""))
        if not z: continue
        for tk in re.split(r"[^A-Za-z']+", str(w)):
            n = norm(tk)
            if len(n) >= 2:
                for base in tg_folds(n):
                    for v in strips(base):
                        tg_g[v].add(z)
    return toda_g, tg_g

# ---------------- candidates ----------------
def branch_xol(n):
    idx = [i for i, c in enumerate(n) if c in "xol"]
    if len(idx) > 8:
        idx = idx[:8]
    outs = set()
    for bits in itertools.product((0, 1), repeat=len(idx)):
        w = list(n)
        for b, i in zip(bits, idx):
            if b: w[i] = {"x": "h", "o": "u", "l": "r"}[n[i]]
        outs.add("".join(w))
    return outs

def contextual(w, short):
    """Safe transforms: regular Truku sound/spelling correspondences."""
    outs = {w}
    subs = [("e", "i"), ("ae", "ay"), ("ai", "ay"), ("ao", "aw"),
            ("ia", "iya"), ("ea", "iya"), ("ui", "uy"), ("ui", "uwi"), ("o", "u"),
            ("wi", "uy")]
    for a, b in subs:
        if w.endswith(a):
            outs.add(w[: -len(a)] + b)
    # palatalization before i is regular — allow even for short tokens (adi > aji)
    outs.add(re.sub(r"d(?=i)", "j", w))
    outs.add(re.sub(r"t(?=i)", "c", w))
    # Pecoraro ao/oa = modern aw/ow/uwa (daolas>dowras, boax>buwax)
    outs.add(w.replace("ao", "aw"))
    outs.add(w.replace("ao", "ow"))
    outs.add(w.replace("oa", "uwa"))
    if not short:
        outs.add(re.sub(r"(?<=[aeiou])u(?=[aei])", "w", w))
        outs.add(re.sub(r"i(?=[aou])", "iy", w))
        outs.add(re.sub(r"e", "i", w))
    return outs

def aggressive(w):
    """Riskier transforms — mapped only with gloss proof."""
    outs = set()
    # Pecoraro writes epenthetic vowels modern Truku drops: kensat>knsat, daxa>dha
    for i, c in enumerate(w[:-1]):
        if c in "eau" and 0 < i < len(w) - 1:
            outs.add(w[:i] + w[i + 1:])
    # Pecoraro writes a palatal glide where modern Truku writes the vowel itself:
    # his y + a/e/o before a consonant is one modern i. NYAQAN 有 is `niqan`
    # (2170x in speech), MNYEQ 居住 is `mniq` (491x), BYEQON 給 is `biqun`,
    # GALYEQ 布料 is `galiq`, QDOLYAQ 逃跑 is `qduriq`, PUNYAQ 火 is `puniq`,
    # XOLYAQ 濕 is `huriq` — which the blind fallback was rendering *huryaq*.
    # It lives here and not in contextual() because the shape it produces is a
    # real word often enough to fool an attestation-only test: SUMYAQ 體蝨 is not
    # `sumiq` 草莓, BASYAQ 暴飲暴食 is not `basiq` (a tree), PUSYAQ 眼屎 is not
    # `pusiq` (a man's name), YAQ 田裡的活 is not `iq` 好. Gloss proof decides.
    outs.add(re.sub(r"y[aeo](?=[^aeiou\W])", "i", w))
    # q/k are inconsistent in Pecoraro (kmpax>qmpah, betak>bitaq) — single swaps
    for i, c in enumerate(w):
        if c == "k":
            outs.add(w[:i] + "q" + w[i + 1:])
        elif c == "q":
            outs.add(w[:i] + "k" + w[i + 1:])
    return outs

def candidates(n):
    """Return {candidate: is_aggressive}."""
    short = len(n) <= 3
    outs = {}
    # two rounds of safe transforms so chains compose (xedao>hedaw>hidaw)
    layer = set()
    for b in branch_xol(n):
        layer |= contextual(b, short)
    safe = set(layer)
    for w in list(layer):
        if len(safe) > 4000:
            break
        safe |= contextual(w, short)
    for c in safe:
        outs.setdefault(c, False)
    if not short:
        for base in list(safe):
            if len(outs) > 8000:
                break
            for c in aggressive(base):
                outs.setdefault(c, True)
    outs.pop(n, None)
    return outs

def gloss_overlap(pec_glosses, omni_glosses):
    """Qualified gloss agreement score; >=2 counts as confirmed.
    Evidence only counts if the whole omnibus gloss is contained in the Pecoraro
    gloss (or vice versa — handles one-char glosses like 日/狗), or the longest
    common substring is >=2 chars AND covers >=20% of the omnibus gloss — a bare
    2-char overlap inside a long unrelated definition is coincidence, not proof
    (raki 哄孩子睡覺… vs laqi 孩子)."""
    best = 0
    for pz in pec_glosses:
        for oz in omni_glosses:
            if not oz: continue
            if oz in pz or pz in oz:
                best = max(best, max(2, min(len(pz), len(oz))))
                continue
            sm = difflib.SequenceMatcher(None, pz, oz, autojunk=False)
            m = sm.find_longest_match(0, len(pz), 0, len(oz))
            if m.size >= 2 and m.size / len(oz) >= 0.2 and m.size > best:
                best = m.size
    return best

def main():
    (tokens, glosses, families, cased, ex_fams, midcap, heads, loan_tokens,
     name_heads) = load_corpus()
    word_raw, word_gloss, sent_best = load_omnibus()
    spoken = load_spoken()
    words_set = set(word_raw)
    attested = words_set | set(sent_best)
    toda_g, tg_g = load_sisters()

    SUFFIXES = ("anay", "ani", "an", "un", "ay", "aw", "i")
    def sister_ev(c):
        """Sister-dialect glosses supporting candidate c — direct, or via the
        candidate's affix-stripped core (>=5 chars) against the stripped index."""
        gs = set(toda_g.get(c, ())) | set(tg_g.get(c, ()))
        for k in (1, 2, 3):
            if len(c) - k >= 5:
                v = c[k:]
                gs |= toda_g.get(v, set()) | tg_g.get(v, set())
        for sfx in SUFFIXES:
            if c.endswith(sfx) and len(c) - len(sfx) >= 5:
                v = c[: -len(sfx)]
                gs |= toda_g.get(v, set()) | tg_g.get(v, set())
                break
        return gs

    # measured correspondence odds (residual-pair counts): o>u and x>h are
    # near-universal (keeping them is the surprise), l usually STAYS l (382 keep
    # vs 70 change) — used only to break ties among equally-glossed candidates
    CH_KEEP = {"o": 0.8, "x": 0.8, "e": 0.5}
    CH_SWAP = {("o", "u"): 0.2, ("x", "h"): 0.2, ("l", "r"): 0.8, ("e", "i"): 0.4}
    def wdist(n, c):
        sm = difflib.SequenceMatcher(None, n, c, autojunk=False)
        cost = 0.0
        for op, i1, i2, j1, j2 in sm.get_opcodes():
            if op == "equal":
                cost += sum(CH_KEEP.get(ch, 0.0) for ch in n[i1:i2])
            elif op == "replace" and (i2 - i1) == (j2 - j1):
                cost += sum(CH_SWAP.get(p, 1.0) for p in zip(n[i1:i2], c[j1:j2]))
            else:
                cost += max(i2 - i1, j2 - j1)
        return cost

    def sister_pick(pool, pec_g, n):
        """pool: {candidate: sister gloss set}. Return the candidate whose
        sister-dialect gloss qualifies (gloss_overlap >= 2) and wins the field —
        gloss score first, then correspondence-weighted closeness to the
        Pecoraro shape n (fold aliases of one sister word would otherwise tie
        forever). A residual exact tie returns None."""
        if not pec_g:
            return None
        hits = []
        for c, gs in pool.items():
            if not gs:
                continue
            g = gloss_overlap(pec_g, gs)
            if g >= 2:
                hits.append((g, c))
        if not hits:
            return None
        hits.sort(key=lambda x: -x[0])
        top = [c for g, c in hits if g == hits[0][0]]
        if len(top) == 1:
            return top[0]
        top.sort(key=lambda c: wdist(n, c))
        return top[0] if wdist(n, top[0]) < wdist(n, top[1]) else None

    result = {}      # token -> record
    review = {}
    unmapped = []    # (token, freq, gloss) with no attested candidate
    tiers = collections.Counter()

    # hand-curated mappings (gloss-verified by a human/LLM pass) win over generated
    manual_path = os.path.join(HERE, "manual_map.json")
    manual = json.load(open(manual_path, encoding="utf-8")) if os.path.exists(manual_path) else {}

    # tier X: LEXICAL SUBSTITUTION, not a respelling. His word is gone from the
    # modern language and a different word carries the meaning (q'nao -> qusul
    # "garlic"). Kept in its own file, and exported separately, because the app
    # must be able to say so on screen: a substituted word renders as
    # MODERN (his original) rather than silently becoming a word he never wrote.
    lex_path = os.path.join(HERE, "lexical_map.json")
    lexical = json.load(open(lex_path, encoding="utf-8")) if os.path.exists(lex_path) else {}
    lexical = {k: v for k, v in lexical.items() if not k.startswith("_")}
    # A null value means the opposite of a substitution: his word is gone and the
    # modern language offers nothing for THIS slot. Sl'xqan is the locative of a
    # verb whose modern replacement (shik) has no locative on record — so the
    # honest answer is silence, and the token is frozen out of every tier and
    # stays green. Without this the rules read it as the skin word and printed
    # srhqan, which is a claim nobody made.
    lex_block = set(k for k, v in lexical.items() if not v)
    lexical = {k: v for k, v in lexical.items() if v}
    OVERRIDE_KEYS.update(lex_block)

    # tier L: per-case adjudication of the C-review queue (gloss-checked one by one)
    llm_path = os.path.join(HERE, "llm_map.json")
    llm = json.load(open(llm_path, encoding="utf-8")) if os.path.exists(llm_path) else {}
    llm = {k: v for k, v in llm.items() if not k.startswith("_")}

    # ---- pass 0c: the loan stratum (tier J) ----
    #
    # It runs above the attestation tiers rather than below them because for this
    # class attestation is actively misleading. Modern standard Truku replaced most
    # of the loan stratum with native coinages (lumak for tobacco, mtgsa for
    # teacher, tluan for table), so when a loan's shape does turn up in the modern
    # corpus it is a homonym — and the more often it turns up, the more confident
    # the wrong answer looked:
    #   TOKE 時計 -> tuki, which is 抵銷 (312x in speech)
    #   DOLI 道理 -> duri, which is 又 (517x)
    #   XAYA 汽車 -> haya, which is 這樣 (129x)
    #   MISO 味噌 -> misu, which is 你 (128x)
    #   BALAS 礫石 -> balas, which is 性交
    # Of the 63 loans the earlier passes claimed, seven landed on a modern word
    # that means the right thing. So a loan may keep an attested spelling only when
    # the GLOSS agrees; otherwise it is romanized by rule and the log says so.
    #
    # The rule is short, because his loan spellings are already internally
    # consistent — across all 122 tagged entries there is not one loan he writes
    # two ways. What they are not is consistent with the rest of the book: he
    # romanizes the source (Japanese o stays o) while his Truku o is modern u.
    # Only that difference is corrected, plus his x for the source h (XINOKI
    # hinoki, XANA hana), his final -e for a vowel modern Truku does not have
    # word-finally (BALE > bali, NABE > nabi, both attested), and the two final
    # glides the modern orthography settled: -ay outnumbers -ai 2129 to 101 and
    # -uy outnumbers -wi 723 to 50. l is left alone — bali 子彈 keeps it, and l>r
    # is guesswork on a word that was never Truku to begin with.
    def loan_rule(t):
        w = plain(t).lower().replace("x", "h").replace("o", "u")
        w = re.sub(r"e$", "i", w)
        w = re.sub(r"ai$", "ay", w)
        return re.sub(r"wi$", "uy", w)

    # Shortest first, so a prefixed form can inherit the base it is built on:
    # KENSAT resolves to knsat on its gloss, and Mkensat must then be Mknsat and
    # not the rule's mkensat. This is tier P's argument confined to the loan set.
    for t in sorted(loan_tokens & set(tokens), key=lambda x: (len(x), x)):
        if t in OVERRIDE_KEYS or t in lexical or len(t) < 2:
            continue
        n = norm(t)
        pec_g = glosses.get(n, set())
        # A hand mapping that CHANGED something was a real decision and still wins
        # (tomato > tmatu). One that only says "leave it alone" was a verdict of
        # "no modern form found", reached before the loans were looked at as a
        # class — nine of them (bolu, kaya, keto, kuli, losok, mkuli, saida,
        # taoke, xama) are exactly the words this pass exists to romanize.
        hand = manual.get(t) or llm.get(t)
        if hand and hand.lower() != plain(t).lower():
            result[t] = {"modern": hand, "tier": "J", "j_how": "hand"}
            tiers["J"] += 1
            continue
        best_g, pick = 0, None
        for c in list(candidates(n)) + [n]:
            if c not in words_set:
                continue
            g = gloss_overlap(pec_g, word_gloss.get(c, set()))
            if g > best_g:
                best_g, pick = g, c
        if best_g >= 2:
            result[t] = {"modern": word_raw.get(pick, pick), "tier": "J", "j_how": "gloss"}
        else:
            base = next((b for b in sorted(loan_tokens, key=len, reverse=True)
                         if b != t and len(b) >= 4 and result.get(b, {}).get("tier") == "J"
                         and t.endswith(b)), None)
            if base:
                result[t] = {"modern": loan_rule(t[:-len(base)]) + result[base]["modern"],
                             "tier": "J", "j_how": "base:" + base}
            else:
                result[t] = {"modern": loan_rule(t), "tier": "J", "j_how": "rule"}
        tiers["J"] += 1

    # OVERRIDE_KEYS freezes the GENERATOR, not the human. The -ui class is in it
    # because the rules cannot choose between ui>uy and ui>uwi from shape alone, so
    # every rule tier is kept off it — but a hand-adjudicated entry in manual_map or
    # llm_map IS the decision the freeze is waiting for, and used to be discarded in
    # silence (batch 46 wrote 13 gloss-verified -ui keys and the rebuild dropped all
    # 13). Human tiers therefore lift the freeze. lex_block does not: a null in
    # lexical_map is itself a human decision, and it means "stay green".
    adjudicated = (set(manual) | set(llm)) - lex_block
    for t in sorted(tokens):
        if (t in OVERRIDE_KEYS and t not in adjudicated) or t in result or len(t) < 2:
            continue
        n = norm(t)
        if not n or len(n) < 2:
            continue
        pec_g = glosses.get(n, set())

        # 0. a lexical substitution outranks everything: no spelling rule can
        # reach it, and no attested-candidate search should be allowed to
        # overrule a decision that was made on the meaning.
        if t in lexical:
            result[t] = {"modern": lexical[t], "tier": "X"}
            tiers["X"] += 1
            continue

        # 0. manual curation wins
        if t in manual:
            result[t] = {"modern": manual[t], "tier": "M"}
            tiers["M"] += 1
            continue

        # 0b. adjudicated review cases
        if t in llm:
            result[t] = {"modern": llm[t], "tier": "L"}
            tiers["L"] += 1
            continue

        # 1. identity
        if n in attested:
            disp = word_raw.get(n, sent_best.get(n, n))
            # When the attested modern spelling matches, keep HIS token so its
            # apostrophes and capitals survive — but strip the diacritics first.
            # norm() ignores them, so däxa/komù/sîda reached this branch and were
            # mapped to themselves, putting ä/ù/î on screen in modern spelling.
            result[t] = {"modern": plain(t) if disp == n else disp, "tier": "id"}
            tiers["id"] += 1
            continue

        # 2-3. generate + filter
        cmap = candidates(n)
        cands = [(c, agg) for c, agg in cmap.items() if c in attested]
        if not cands:
            # no Truku attestation — triangulate via Toda/Tgdaya cognates.
            # Identity (keep the Pecoraro spelling) competes as a candidate too.
            pool = {c: sister_ev(c) for c in cmap}
            pool[n] = sister_ev(n)
            pick = sister_pick(pool, pec_g, n)
            if pick is not None:
                disp = word_raw.get(pick, sent_best.get(pick, pick))
                result[t] = {"modern": t if disp == n else disp, "tier": "T"}
                tiers["T"] += 1
            else:
                tiers["none"] += 1
                unmapped.append((t, tokens[t], sorted(pec_g)[:1]))
            continue

        # 4. score
        scored = []
        for c, agg in sorted(cands):
            in_words = c in words_set
            g = gloss_overlap(pec_g, word_gloss.get(c, set())) if in_words else 0
            dist = sum(1 for a, b in zip(n, c) if a != b) + abs(len(n) - len(c))
            scored.append({"cand": c, "words": in_words, "gloss": g, "dist": dist, "agg": agg})
        # safety outranks gloss unless the gloss evidence is decisive: an
        # aggressive candidate must beat every safe candidate by >=2 gloss points
        safe_best_gloss = max([s["gloss"] for s in scored if not s["agg"]], default=-1)
        def rank(s):
            decisive = s["agg"] and s["gloss"] >= 2 and s["gloss"] >= safe_best_gloss + 2
            return (-s["gloss"] if (not s["agg"] or decisive) else 0,
                    s["agg"] and not decisive, s["agg"], not s["words"], s["dist"])
        scored.sort(key=rank)
        best = scored[0]
        disp = word_raw.get(best["cand"], sent_best.get(best["cand"], best["cand"]))
        safe = [s for s in scored if not s["agg"]]

        # 5. tier — aggressive candidates need gloss proof
        if best["words"] and best["gloss"] >= 2 and (
            len(scored) == 1 or best["gloss"] > scored[1]["gloss"] or scored[1]["dist"] > best["dist"]
        ):
            result[t] = {"modern": disp, "tier": "A"}
            tiers["A"] += 1
        elif len(safe) == 1:
            c = safe[0]
            d2 = word_raw.get(c["cand"], sent_best.get(c["cand"], c["cand"]))
            result[t] = {"modern": d2, "tier": "B"}
            tiers["B"] += 1
        elif safe:
            # ambiguous among safe candidates: prefer the plain-rules output
            rules_out = re.sub(r"[xol]", lambda m: {"x": "h", "o": "u", "l": "r"}[m.group(0)], n)
            hit = next((s for s in safe if s["cand"] == rules_out), None)
            if hit is not None:
                d2 = word_raw.get(rules_out, sent_best.get(rules_out, rules_out))
                result[t] = {"modern": d2, "tier": "B"}
                tiers["B-rules"] += 1
            else:
                # ambiguous among safe candidates — let a sister-dialect cognate
                # gloss break the tie (conservative: safe candidates only)
                pool = {s["cand"]: sister_ev(s["cand"]) for s in safe}
                pick = sister_pick(pool, pec_g, n)
                if pick is not None:
                    d2 = word_raw.get(pick, sent_best.get(pick, pick))
                    result[t] = {"modern": d2, "tier": "T"}
                    tiers["T"] += 1
                else:
                    review[t] = {"freq": tokens[t], "pec_gloss": sorted(pec_g)[:2], "cands": scored[:5]}
                    tiers["C-review"] += 1
        else:
            # only aggressive candidates: a sister-dialect gloss match counts as
            # the required gloss proof
            pool = {s["cand"]: sister_ev(s["cand"]) for s in scored}
            pick = sister_pick(pool, pec_g, n)
            if pick is not None:
                d2 = word_raw.get(pick, sent_best.get(pick, pick))
                result[t] = {"modern": d2, "tier": "T"}
                tiers["T"] += 1
            else:
                review[t] = {"freq": tokens[t], "pec_gloss": sorted(pec_g)[:2], "cands": scored[:5]}
                tiers["C-review"] += 1

    # ---- pass 2: root-consistency projection (tier P) ----
    # A resolved family member (tiers id/M/A/B) fixes the Pecoraro->modern
    # correspondence for its stem; unresolved members of the same entry family
    # (hw + subs + paradigm forms — NOT example tokens, which mix in unrelated
    # words) inherit it: candidate = convert(prefix) + modern_stem + convert(
    # suffix), allowing an m/n-type infix after the stem's first consonant.
    # Affix conversion is limited to the near-universal correspondences
    # (o>u, x>h, final -ai/-ao/-e). Projected forms are mostly unattested by
    # definition — the point is inheriting a verified stem, and protecting
    # already-modern derivatives from the blanket character rules.
    AFFIX_END = [("ai", "ay"), ("ao", "aw"), ("e", "i")]
    def affix_convert(a, final):
        a = a.replace("o", "u").replace("x", "h")
        if final:
            for src, dst in AFFIX_END:
                if a.endswith(src):
                    a = a[: -len(src)] + dst
                    break
        return a

    INFIXES = ("mn", "um", "nm", "m", "n")

    # Truku doesn't write the schwa, so a root loses its first vowel the moment
    # anything is prefixed to it: GAMIL 根 is the root, but "the place where it
    # took root" is Tgmilan, not *Tgamilan. Testing the stem by literal
    # containment therefore misses a word's own conjugates — and tier R, reaching
    # the token with no family to answer to, guessed tgmiran and flipped an l the
    # root plainly keeps. Every resolved stem is offered in both shapes; the
    # syncopated one only counts when both sides syncopate the same way, which is
    # what makes it a correspondence rather than a second guess.
    def stem_forms(sp, sm):
        out = [(sp, sm)]
        mp = re.match(r"^([^aeiou'])[aeiou](?=[^aeiou])", sp)
        mm_ = re.match(r"^([^aeiou'])[aeiou](?=[^aeiou])", sm)
        if mp and mm_ and len(sp) >= 4 and len(sm) >= 4:
            out.append((sp[0] + sp[2:], sm[0] + sm[2:]))
        return out

    proposals = collections.defaultdict(set)     # token -> candidate modern forms
    for fam in families:
        resolved = []
        for m in sorted(set(fam)):
            rec = result.get(m)
            if not rec:
                continue
            sp = norm(m)
            if len(sp) >= 3:
                resolved.extend(stem_forms(sp, rec["modern"].lower()))
        if not resolved:
            continue
        resolved.sort(key=lambda x: (-len(x[0]), x[0], x[1]))
        for t in sorted(set(fam)):
            if t in result or t in OVERRIDE_KEYS or len(t) < 2:
                continue
            n = norm(t)
            if len(n) < 3:
                continue
            for sp, sm in resolved:
                hit = None                       # (start, matched_len, modern_core)
                i = n.find(sp)
                if i >= 0:
                    hit = (i, len(sp), sm)
                else:
                    for inf in INFIXES:
                        i = n.find(sp[0] + inf + sp[1:])
                        if i >= 0:
                            hit = (i, len(sp) + len(inf), sm[0] + inf + sm[1:])
                            break
                if hit is None:
                    continue
                i, L, core = hit
                pre, suf = n[:i], n[i + L:]
                if len(pre) > 5 or len(suf) > 4:
                    continue
                proposals[t].add(affix_convert(pre, False) + core + affix_convert(suf, True))
                break                            # longest matching stem wins per family
    proj_att = proj_ambig = 0
    for t, cs in sorted(proposals.items()):
        if t in result:
            continue
        if len(cs) != 1:
            proj_ambig += 1
            continue
        cand = next(iter(cs))
        if cand in attested:
            disp = word_raw.get(cand, sent_best.get(cand, cand))
            proj_att += 1
        else:
            disp = t if cand == norm(t) else cand
        result[t] = {"modern": disp, "tier": "P"}
        tiers["P"] += 1
    tiers["none"] -= tiers["P"]
    unmapped = [u for u in unmapped if u[0] not in result]

    # ---- pass 2b: relative inheritance (tier R) ----
    # Everything above tests a WHOLE word against the omnibus. Truku is heavily
    # affixing, so a regularly derived form of a perfectly well attested root
    # falls straight through: nduk is nowhere in the omnibus, but mduk 關（門、窗）
    # and mnduk 曾關門 are right there. Tier P covers this only when a relative
    # happens to sit in the SAME Pecoraro entry; this pass goes to the omnibus
    # for the relative instead. (load_sisters() has done affix-stripped core
    # matching for Toda/Tgdaya all along — it was never done for Truku itself.)
    #
    # Guards, because a shared stem is much weaker evidence than a shared word:
    #   - core >= 3 letters, and only SAFE candidate transforms on the core
    #   - the core must be shared by >=2 distinct glossed omnibus words, so it is
    #     a real root and not an accident of stripping
    #   - exactly one modern reading may survive; two readings means don't guess
    #   - if both sides carry a Chinese gloss and they don't overlap at all,
    #     veto — that is a false friend riding a coincidental stem (raki/laqi)
    # Affixes convert by the near-universal rules only (o>u, x>h); l>r can only
    # happen inside a core that an attested modern word actually spells with r,
    # so this pass cannot make the keep-l mistake pass 3 exists to undo.
    R_MINC = 3
    R_PREF = ("mnp", "dmp", "mpp", "mn", "kn", "pn", "tn", "gn", "sm", "nk", "pk",
              "sk", "tk", "mk", "dm", "sn", "ps", "pg", "km", "gm", "tm", "mp",
              "np", "p", "s", "t", "n", "k", "m", "g", "d", "b", "q", "h", "c", "j")
    R_SUFF = ("anay", "ani", "an", "un", "ay", "aw", "on", "i")

    def peel(w, dered):
        """(prefix, infix, core, suffix) splits of w with a core >= R_MINC.
        The infix is tracked apart from the prefix because it belongs after the
        core's first consonant: xngloq is x+n+gloq, so it rebuilds as h+n+gluq,
        not n+hgluq."""
        outs = set()
        for p in [""] + [p for p in R_PREF if w.startswith(p)]:
            rest = w[len(p):]
            infd = [("", rest)]
            for inf in INFIXES:
                if len(rest) > len(inf) + 1 and rest[1:1 + len(inf)] == inf:
                    infd.append((inf, rest[0] + rest[1 + len(inf):]))
            for inf, r in infd:
                for sfx in ("",) + R_SUFF:
                    if sfx and not r.endswith(sfx):
                        continue
                    core = r[: len(r) - len(sfx)] if sfx else r
                    if dered and len(core) > R_MINC and core[0] == core[1]:
                        outs.add((p, inf, core[1:], sfx))     # llukus -> lukus
                    if len(core) >= R_MINC:
                        outs.add((p, inf, core, sfx))
        return outs

    core_index = collections.defaultdict(set)      # core -> attested modern words
    for w in words_set:
        for _, _, core, _ in peel(w, True):
            core_index[core].add(w)

    r_ambig = r_veto = 0
    for t in sorted(tokens):
        if t in result or t in OVERRIDE_KEYS:
            continue
        n = norm(t)
        if len(n) < 3:
            continue
        outs = collections.defaultdict(set)        # modern form -> supporting words
        for pre, inf, core, sfx in peel(n, False):
            cands = {c for c, agg in candidates(core).items() if not agg} | {core}
            for cc in cands:
                sup = core_index.get(cc)
                if sup and len(sup) >= 2:
                    body = (cc[0] + inf + cc[1:]) if inf else cc
                    outs[affix_convert(pre, False) + body + affix_convert(sfx, True)] |= sup
        if not outs:
            continue
        if len(outs) > 1:
            r_ambig += 1
            continue
        pick = next(iter(outs))
        sup = outs[pick]
        # Gloss veto. gloss_overlap() is a PROMOTION test — it wants a contiguous
        # substring, because a 2-char run inside a long definition is coincidence.
        # Here the job is the opposite (reject a false friend), and the substring
        # test is far too harsh for it: nduk 門關閉的 and mduk 關門窗 plainly agree,
        # yet share no run longer than one character because the order differs.
        # So: veto only when the two glosses have no character in common at all.
        pec_g = glosses.get(n, set())
        omni_g = set().union(*(word_gloss.get(w, set()) for w in sup))
        if pec_g and omni_g and not (set("".join(pec_g)) & set("".join(omni_g))):
            r_veto += 1
            continue
        result[t] = {"modern": word_raw.get(pick, t if pick == n else pick), "tier": "R"}
        tiers["R"] += 1
        review.pop(t, None)
    tiers["none"] -= sum(1 for u in unmapped if u[0] in result)
    unmapped = [u for u in unmapped if u[0] not in result]

    # ---- pass 3: keep-l guard (tier KL) ----
    # The app fallback applies o>u, l>r, x>h blindly. l>r is the "expensive"
    # rule — l usually stays l in Truku — and it wrongly corrupts derived forms
    # of l-keeping roots whose inflected surface form isn't itself in the omnibus
    # (llukus>rrukus though root lukus stays lukus; l'alang>r'arang though alang
    # stays alang). A whole-token dictionary check can't catch these because only
    # the ROOT is attested. So: strip affixes/reduplication; if a stripped root's
    # keep-l form is attested in the Words sheet while its l>r form is not, freeze
    # the token to its keep-l spelling (o>u, x>h, l untouched) instead of char-ruling.
    KL_PREF = ("mn", "kn", "pn", "tn", "gn", "sm", "nk", "pk", "sk", "tk", "mk",
               "dm", "sn", "ps", "pg", "km", "gm", "tm", "p", "s", "t", "n", "k",
               "m", "g", "d", "b", "q", "l", "x")
    def keep_l(s):     # modernize but leave l alone (matches app fallback minus l>r)
        return s.replace("o", "u").replace("x", "h")
    def l_to_r(s):
        return s.replace("o", "u").replace("l", "r").replace("x", "h")
    def kl_cores(n):
        outs = set()
        for p in KL_PREF:
            if n.startswith(p) and len(n) - len(p) >= 4:
                outs.add(n[len(p):].lstrip("'"))
        if "'" in n:
            outs.add(n.split("'", 1)[1].lstrip("'"))
            outs.add(n.rsplit("'", 1)[1])
        if len(n) >= 2 and n[0] == n[1]:              # de-reduplicate llukus>lukus
            outs.add(n[1:])
        return {c for c in outs if len(c) >= 4}
    kl = 0
    for t in sorted(tokens):
        if t in result or t in OVERRIDE_KEYS or len(t) < 2:
            continue
        n = norm(t)
        if "l" not in n or l_to_r(n) == keep_l(n):     # no l that l>r would change
            continue
        for c in kl_cores(n):
            if keep_l(c) != l_to_r(c) and keep_l(c) in words_set and l_to_r(c) not in attested:
                result[t] = {"modern": keep_l(t), "tier": "KL"}
                tiers["KL"] += 1
                review.pop(t, None)
                kl += 1
                break
    tiers["none"] -= sum(1 for u in unmapped if u[0] in result)
    unmapped = [u for u in unmapped if u[0] not in result]

    # ---- pass 3c: attestation in running Truku speech (tier S) ----
    # Same claim as tiers A/B — "this exact word exists in modern Truku" — but
    # asked of a 277k-token body of transcribed speech rather than of a
    # dictionary. Candidates are the rule-consistent readings of his token (each
    # o/l/x independently kept or converted, plus the near-universal final
    # -ai>-ay / -ao>-aw / -e>-i), and exactly one of them must be in the corpus
    # at least twice: once is as likely to be an ASR slip as a word.
    S_END = [("ai", "ay"), ("ae", "ay"), ("ao", "aw"), ("e", "i")]
    def rule_readings(n):
        outs = set()
        for keep in itertools.product(*[("olx".find(c) >= 0 and (c, {"o": "u", "l": "r", "x": "h"}[c]) or (c,))
                                        for c in n]):
            w = "".join(keep)
            outs.add(w)
            for src, dst in S_END:
                if w.endswith(src):
                    outs.add(w[: -len(src)] + dst)
        return outs
    s_log = []
    for t in sorted(tokens):
        if t in result or t in OVERRIDE_KEYS:
            continue
        n = norm(t)
        if len(n) < 4:
            continue
        hits = sorted(c for c in rule_readings(n) if spoken.get(c, 0) >= 2)
        if len(hits) != 1 or hits[0] in attested:
            continue
        # A corpus hit that flips an l is only believable if the keep-l reading of
        # the root is NOT itself a modern word: mk'alang matched karang (crab) in
        # transcribed speech, but his word is built on alang (village).
        if "l" in n and l_to_r(n) == hits[0]:
            if any(keep_l(c) in words_set for c in kl_cores(n) | {n}):
                continue
        result[t] = {"modern": hits[0], "tier": "S", "spoken": spoken[hits[0]]}
        tiers["S"] += 1
        review.pop(t, None)
        s_log.append((t, hits[0], spoken[hits[0]], tokens[t]))
    tiers["none"] -= sum(1 for u in unmapped if u[0] in result)
    unmapped = [u for u in unmapped if u[0] not in result]

    # ---- pass 3b: proper names (tier N) ----
    # "Sapah Sibar u, ana manu ida stbaku kana da!" — Sibal is a man, and the
    # blind l>r rule renamed him. A name is not a common noun: nothing attests it
    # and no tier above will ever reach it, so it falls to the char rules, which
    # is the one population where they are guaranteed to be guessing. Test: the
    # token is capitalized in the middle of one of his sentences (only a proper
    # noun is) and is never written lowercase anywhere in the book. Those keep
    # their l; o>u, x>h and the final -ai/-ao conversions are near-universal and
    # still apply, so Pisao becomes Pisaw and Labai Labay while Sibal stays Sibal.
    #
    # Runs after tier S on purpose, and the `t in result` guard below is what keeps
    # it there: attestation outranks the freeze. The community really does write
    # KULAS as kuras (24x), LABAI rabay (42x), LIBIç ribix (11x) — those are tier S
    # and must stay r. The freeze is for the names no corpus has an opinion about.
    n_log = []
    for t in sorted(set(midcap) | name_heads):
        if t in result or t in OVERRIDE_KEYS or len(norm(t)) < 3:
            continue
        if t not in name_heads:
            # "Never lowercase anywhere" is one stray keystroke away from failing.
            # Wilang is Wilang 9 times mid-sentence and WILANG once as a headword,
            # and wilang exactly once — and that one slip vetoed the man. So the
            # veto now needs the lowercase reading to be more than a slip: mid-
            # sentence capitals must still be at least 60% of every occurrence.
            # Measured: this admits 5 tokens and all 5 are proper nouns — Wilang and
            # Dloan (men), Taolan (a neighbour), Tagaxan (a place he climbs to) and
            # Taiwan. Dropping the 60% and asking only that capitals outnumber
            # lowercase admits 142, led by ini, ana, adi and malu, which are
            # capitalized because they start his sentences; that gate is useless.
            # A tag seed needs none of this: `name (m)` is not a statistic.
            low = sum(v for s, v in cased.get(t, {}).items() if s[:1].islower())
            if low and midcap[t] < 0.6 * tokens[t]:
                continue
        # The ending conversion has to be asked BEFORE keep_l, not after. keep_l is
        # o>u, so by the time it returned, -ao was already -au and endswith("ao")
        # could never match: the documented "Pisao becomes Pisaw" never fired, and
        # the tier quietly emitted -au. It went unnoticed because tier S owns every
        # attested -aw name (asaw 91x, tadaw 71x, umaw 66x) and the only -au tier N
        # ever reached was beau, unattested either way. The tag seed makes it live —
        # amai, dawai, masai, tilae are all reached now — so fix the order.
        m = plain(t)
        for src, dst in S_END[:3]:
            if m.lower().endswith(src):
                m = m[: -len(src)] + dst
                break
        m = keep_l(m)
        result[t] = {"modern": m, "tier": "N"}
        tiers["N"] += 1
        review.pop(t, None)
        n_log.append((t, m, midcap[t], tokens[t]))
    tiers["none"] -= sum(1 for u in unmapped if u[0] in result)
    unmapped = [u for u in unmapped if u[0] not in result]

    # ---- pass 4: morphology over an already-solved base (tier D) ----
    # Lowking Nowbucyang, 太魯閣語構詞法研究 (Word Formation in Truku, 2008) §3.4:
    # Truku reduplication is CV- or CVCV-. Truku does not write the schwa, so CV-
    # surfaces in the orthography as a DOUBLED INITIAL CONSONANT — hmadan
    # "clear a field" → hhmadan "many of them clearing". Two more processes behave
    # the same way for our purposes: the mn-/n- AF preterite (mhmadan → mnhmadan)
    # and the collective d- on a personal name (Aman → dAman, "Aman and his
    # group"). None of the three makes a new lexeme, so the modern spelling is
    # just (his affix) + the modern spelling of the base.
    #
    # Every other tier tests the token as a WHOLE against the omnibus, and tier R
    # peels affixes off the Pecoraro side but never de-reduplicates it — peel() is
    # called with the reduplication flag on the modern side only. That is exactly
    # why these fell through to the blind character rules, where l>r then
    # corrupted them: llisao reached the screen as *rrisau* for rrisaw, xxei as
    # *hhei* for hhiyi, nk'la as *nk'ra* for a root that keeps its l.
    #
    # Runs last, so a base fixed by ANY earlier tier counts — except tier X, whose
    # "modern" is a different word: a derived form of q'nao must not quietly
    # become a derived form of qusul.
    VOWELS = set("aeiouàáâäèéêë"
                 "ìíîïòóôöùúûü")
    def is_cons(c):
        return c.isalpha() and c not in VOWELS
    def based(w):
        r = result.get(w)
        return r["modern"] if r and r["tier"] != "X" else None
    d_rules = collections.Counter()
    d_log = []
    for t in sorted(tokens):
        if t in result or t in OVERRIDE_KEYS or len(t) < 3:
            continue
        got = None
        # CV-: double the MODERN initial, which is not always his own —
        # kksaxol is built on ksaxol = qsahur, so it is qqsahur, not *kksahur.
        if is_cons(t[0]) and t[0] == t[1]:
            m = based(t[1:])
            if m and m[:1].isalpha():
                got = ("CV-", t[1:], m[0] + m)
        if not got and len(t) > 4 and t[0:2] == t[2:4] and is_cons(t[0]):
            m = based(t[2:])
            if m and len(m) > 1 and m[0].isalpha() and m[1].isalpha():
                got = ("CVCV-", t[2:], m[0:2] + m)
        if not got:
            for pre in ("mn", "n"):
                if not t.startswith(pre) or len(t) <= len(pre) + 2:
                    continue
                rest = t[len(pre):]
                m = based(rest)
                if m:
                    got = (pre + "-", rest, pre + m)
                    break
                # his AF base: mali → nali. Only strip the m if the modern form
                # still carries one, or the strip is cutting a different letter.
                m = based("m" + rest)
                if m and m[:1] == "m":
                    got = (pre + "-", "m" + rest, pre + m[1:])
                    break
        if not got and t[0] == "d" and len(t) > 3:
            forms = cased.get(t[1:], {})
            caps = sum(v for k, v in forms.items() if k[:1].isupper())
            m = based(t[1:])
            if m and caps and caps == sum(forms.values()):
                got = ("d-", t[1:], "d" + m)
        if got:
            rule, base, mod = got
            result[t] = {"modern": mod, "tier": "D", "d_rule": rule, "d_base": base}
            tiers["D"] += 1
            d_rules[rule] += 1
            d_log.append((t, mod, rule, base, result[base]["tier"]))
            review.pop(t, None)
    tiers["none"] -= sum(1 for u in unmapped if u[0] in result)
    unmapped = [u for u in unmapped if u[0] not in result]

    # ---- pass 5: projection into his own example sentences (tier E) ----
    # Tier P deliberately refuses example tokens, because a sentence is mostly
    # other people's words. But that also shut the door on a word's own family:
    # XEBONG is resolved to hibung on the headword, Mxebong to mhibung on the
    # sub-form, and kxebong — which occurs nowhere but the single sentence under
    # that entry — stayed green and went on screen as khebung. It is not a
    # borderline case; it is the same word.
    #
    # The claim here is narrower than tier P's. A sentence token qualifies only if
    # it CONTAINS a stem the same entry has already resolved, so ka, so and ini
    # can never match, and one ambiguous candidate disqualifies the token. On the
    # 5,438 sentences that leaves 148 words, none of them ambiguous, and 92 of
    # them are words the blind rules are getting wrong today — nearly all through
    # l>r on a keep-l root (msnugul as msnugur, gnluban as gnruban) or h>x on a
    # keep-x one (mtgimax as mtgimah).
    e_log = []
    ex_proposals = collections.defaultdict(set)
    for fam, ext in ex_fams:
        stems = []
        for m in sorted(set(fam)):
            rec = result.get(m)
            if rec and rec["tier"] != "X" and len(norm(m)) >= 3:
                stems.extend(stem_forms(norm(m), rec["modern"].lower()))
        if not stems:
            continue
        stems = sorted(set(stems), key=lambda x: (-len(x[0]), x[0], x[1]))
        for t in sorted(set(ext)):
            if t in result or t in OVERRIDE_KEYS:
                continue
            n = norm(t)
            if len(n) < 3:
                continue
            for sp, sm in stems:
                hit = None
                i = n.find(sp)
                if i >= 0:
                    hit = (i, len(sp), sm)
                else:
                    for inf in INFIXES:
                        i = n.find(sp[0] + inf + sp[1:])
                        if i >= 0:
                            hit = (i, len(sp) + len(inf), sm[0] + inf + sm[1:])
                            break
                if hit is None:
                    continue
                i, L, core = hit
                pre, suf = n[:i], n[i + L:]
                if len(pre) > 5 or len(suf) > 4:
                    continue
                ex_proposals[t].add((affix_convert(pre, False) + core +
                                     affix_convert(suf, True), sp, sm))
                break
    for t, cs in sorted(ex_proposals.items()):
        if t in result:
            continue
        if len(set(c[0] for c in cs)) != 1:
            continue
        cand, sp, sm = sorted(cs)[0]
        disp = word_raw.get(cand, cand) if cand in attested else cand
        result[t] = {"modern": disp, "tier": "E", "e_stem": sp, "e_modern": sm}
        tiers["E"] += 1
        e_log.append((t, disp, sp, sm, "attested" if cand in attested else "derived"))
        review.pop(t, None)
    tiers["none"] -= sum(1 for u in unmapped if u[0] in result)
    unmapped = [u for u in unmapped if u[0] not in result]

    # ---- pass 6: root projection across entries (tier G) ----
    # Tier E only sees the entry a token stands in, and words don't respect that
    # boundary: `mptgamil` occurs once, in a sentence under GABAL 拔, so nothing in
    # its own entry could tell it that GAMIL 根 is right there resolved and keeps
    # its l. It reached the screen as mptgamir. So the same containment test is run
    # once more against the dictionary's ROOTS rather than the entry's family.
    #
    # Being global, it is held to a higher bar than tier E: the root must be one
    # the corpus actually vouches for (no projected or inherited tier may seed
    # another projection), at least 4 characters so a coincidental substring
    # cannot qualify, and the root's own reading must be unique — two roots that
    # normalize to the same shape disqualify each other. One ambiguous candidate
    # still kills the token.
    #
    # Two further guards, both learned from the first run of this pass:
    #  * The stem pair must be a letter-for-letter CORRESPONDENCE. Pecoraro's MALO
    #    also surfaces as `nalu`, and reading that as a spelling rule projected
    #    `mpanalu` to *mpamalu — n/m there is the AF prefix alternating, not an
    #    orthography. Only the attested letter swaps may differ.
    #  * Tier G only speaks when it has something to say. A root that the blind
    #    rules already convert correctly adds no evidence, and a 4-letter substring
    #    match is cheap: `banasi` contains `nasi`, `kliban` contains the name
    #    `Iban`. Where the projection agrees with the fallback the token is left
    #    green, which is the honest colour for "nothing checked this".
    #  * Only example-sentence tokens are eligible, and no root may seed a token
    #    that Pecoraro filed under a root of its own. `Kliban` is a sub-form of
    #    KALIP 剪, so the fact that it ends in the name `Iban` is nothing; tiers
    #    P/R/KL own that word and may leave it green if they can't settle it.
    #  * A name is never a stem. Tier N words are people, and people turn up
    #    inside other words by accident only.
    G_SEED = {"id", "M", "L", "A", "B", "B-rules", "T", "S"}
    # `families` only collects an entry with more than one form, so a lone headword
    # like QALIP is not in it — and QALIP is a headword Pecoraro himself files with
    # KALIP. Add the heads back or tier G walks straight into it.
    family_tokens = set(t for fam in families for t in fam) | set(heads)
    G_SWAP = {("o", "u"), ("l", "r"), ("x", "h"), ("k", "q"), ("q", "k"),
              ("d", "j"), ("t", "c"), ("e", "i"), ("ç", "x")}
    def corresponds(sp, sm):
        if len(sp) != len(sm):
            return False
        return all(a == b or (a, b) in G_SWAP for a, b in zip(sp, sm))
    root_stems = collections.defaultdict(set)
    for h in sorted(set(heads)):
        rec = result.get(h)
        if not rec or rec["tier"] not in G_SEED or len(norm(h)) < 4:
            continue
        for sp, sm in stem_forms(norm(h), rec["modern"].lower()):
            if len(sp) >= 4 and corresponds(sp, sm):
                root_stems[sp].add(sm)
    root_stems = {sp: next(iter(ms)) for sp, ms in root_stems.items() if len(ms) == 1}
    g_stems = sorted(root_stems.items(), key=lambda x: (-len(x[0]), x[0], x[1]))
    # Local root wins. `qalip` sits in a sentence under KALIP 剪 and is the same
    # word as the headword under the very q/k swap this pass allows — so it is not
    # free to go and inherit from QALI 話 across the book. Same length and a
    # letter-for-letter correspondence is a deliberately tight test.
    g_local = set()
    for fam, ext in ex_fams:
        fn = set(norm(f) for f in fam if len(norm(f)) >= 4)
        for t in sorted(set(ext)):
            n = norm(t)
            if any(corresponds(n, f) for f in fn):
                g_local.add(t)
    g_log = []
    for t in sorted(tokens):
        if t in result or t in OVERRIDE_KEYS or t in family_tokens or t in g_local:
            continue
        n = norm(t)
        if len(n) < 4:
            continue
        cands = set()
        for sp, sm in g_stems:
            hit = None
            i = n.find(sp)
            if i >= 0:
                hit = (i, len(sp), sm)
            else:
                for inf in INFIXES:
                    i = n.find(sp[0] + inf + sp[1:])
                    if i >= 0:
                        hit = (i, len(sp) + len(inf), sm[0] + inf + sm[1:])
                        break
            if hit is None:
                continue
            i, L, core = hit
            pre, suf = n[:i], n[i + L:]
            if len(pre) > 3 or len(suf) > 3 or len(pre) + len(suf) > 5:
                continue
            cands.add((affix_convert(pre, False) + core + affix_convert(suf, True), sp, sm))
        if len(set(c[0] for c in cands)) != 1:
            continue
        cand, sp, sm = sorted(cands)[0]
        if cand == l_to_r(n):
            continue    # the blind fallback already lands here; claim nothing
        disp = word_raw.get(cand, cand) if cand in attested else cand
        result[t] = {"modern": disp, "tier": "G", "g_stem": sp, "g_modern": sm}
        tiers["G"] += 1
        review.pop(t, None)
        g_log.append((t, disp, sp, sm, tokens[t], "attested" if cand in attested else "derived"))
    tiers["none"] -= sum(1 for u in unmapped if u[0] in result)
    unmapped = [u for u in unmapped if u[0] not in result]

    # Nothing leaves this generator carrying his diacritics. A modern spelling is
    # written in the modern alphabet, so ç becomes x and the vowel marks drop —
    # whatever tier produced it. The identity tier used to return his raw token
    # (däxa -> däxa) and tier P inherited a marked stem (msueq -> msüeq); the app
    # cannot repair either, because a map hit short-circuits its character rules.
    for rec in result.values():
        pl = plain(rec["modern"])
        if pl != rec["modern"]:
            rec["diacritics_dropped"] = rec["modern"]
            rec["modern"] = pl

    # Nor does anything leave carrying his elision marks. Pecoraro writes ' and "
    # where he heard a schwa; the modern orthography writes nothing at all, on the
    # same principle that gives Truku dxgal, mkla, tmlung. Measured across the
    # three modern corpora: ONE token in 39,889 contains an apostrophe, and it is
    # a typo (`kiya hug'`). The marks were surviving because the identity tier
    # hands back HIS token whenever the attested spelling matches (`d'xgal` norms
    # to the attested `dxgal`, so disp == n and he was echoed verbatim), and the
    # projection tiers then inherited a marked stem: 312 types / 921 occurrences
    # on screen, `d'xgal` 土地 x49, `mk'la` 知道 x39, `'bi` x15. Stripping is not a
    # guess — for 111 of the 120 identity cases the bare form is the attested
    # modern word, usually a very common one (dxgal 662 in speech, mkla 671,
    # bi 6292). It runs after the tiers rather than inside plain(), which is also
    # used to compare against his token where the mark still has to be there.
    for rec in result.values():
        bare = ELISION_RE.sub("", rec["modern"])
        if bare != rec["modern"]:
            rec["elision_dropped"] = rec["modern"]
            rec["modern"] = bare

    # A modern spelling with no vowel in it is not a word. Pecoraro writes the schwa
    # he heard — with ö, with ', with " — and the tiers that peel his marks off can
    # peel away the only vowel a token had: SB'LÖS "fade, sans saveur" has the
    # variant SB'L'S in its own tag, and tier R turned it into *sbls, SK'L'T into
    # *skrt. Both are unpronounceable, and neither is what the corpus says (sblus
    # 不鹹;不甜, msblus 沒有味道 — his gloss exactly). The omnibus's only vowelless
    # entries are seven abbreviations (mk, wy, sk, msn...), so this costs nothing and
    # the token goes back to honest green rather than on screen as a consonant run.
    devowelled = []
    for t in [t for t, rec in result.items()
              if rec["tier"] not in ("X",)
              and not set("aeiou") & set(rec["modern"].lower())]:
        devowelled.append((t, result[t]["modern"], result[t]["tier"]))
        del result[t]
        unmapped.append((t, tokens[t], []))
        tiers["none"] += 1
    if devowelled:
        print("no-vowel gate: %d refused — %s" % (
            len(devowelled), ", ".join("%s->%s(%s)" % d for d in sorted(devowelled))))

    # ---- pass 7: his own elision-mark variants (tier V) ----
    # Pecoraro writes the schwa he heard with ' or ", and he is not consistent about
    # it WITHIN one word: TNQDO's tag reads "(= R. = L'QDO ?)" while the entry it
    # points at is LQDO. Those are two different map keys — wordKey() folds the two
    # marks together but does not remove them — so the marked twin never reached the
    # map and went on screen in char-rule green, RQDU beside its own brown RQDUG.
    # He does this most in the bracketed variant tags where he is explicitly listing
    # spellings of one word: (TNG'I - T'NGI), (X'GUT ?), (PG'DGIT), (BQ'LI).
    #
    # So an unmapped token inherits from its elision-free twin. Three guards, and
    # they are the whole safety of the tier:
    #  - the twin's reading must be UNIQUE. Folding the marks blindly is not safe:
    #    seven bare shapes carry keys that disagree (b'xgan/bxgan, kn'qan/knqan,
    #    mq'qan/mqqan, p'lapa/plapa, wa'lo/walo), because for those he is writing
    #    two different words, not one word two ways. Requiring one value skips all
    #    of them and leaves them honestly green.
    #  - never a tier X key. A lexical substitution has to declare itself on screen,
    #    and q'nao / sl'xeq / sml'xeq / t'bako all have a mark-free twin: inheriting
    #    the value would print a bare brown QUSUL and bypass the (Q'NAO) disclosure.
    #  - never a lex_block token. Those are green on purpose — we looked and there is
    #    no modern form for that slot — so a twin must not answer for them.
    v_log = []
    by_bare = collections.defaultdict(dict)
    for k, rec in result.items():
        by_bare[ELISION_RE.sub("", k)][k] = rec
    for t, o, _g in unmapped:
        b = ELISION_RE.sub("", t)
        if b == t and len(t) < 3:
            continue
        twins = by_bare.get(b)
        if not twins or t in lexical or t in lex_block:
            continue
        if any(r["tier"] == "X" for r in twins.values()):
            continue
        vals = set(r["modern"] for r in twins.values())
        if len(vals) != 1:
            continue
        src = sorted(twins)[0]
        result[t] = {"modern": twins[src]["modern"], "tier": "V", "v_twin": src}
        tiers["V"] += 1
        review.pop(t, None)
        v_log.append((t, twins[src]["modern"], src, twins[src]["tier"], o))
    tiers["none"] -= sum(1 for u in unmapped if u[0] in result)
    unmapped = [u for u in unmapped if u[0] not in result]

    # 7b. Same evidence, the other direction: a hand-verified twin beats a machine
    # one. The pass above only reaches tokens NOTHING mapped, so a token whose other
    # elision spelling was settled by hand can still be sitting on a rules value, and
    # then his two marks print two different words: `mg'li` came out *mgli beside the
    # verified `mg'li"` → mgrig 跳舞. Measured over the finished map: 16 twin groups
    # hold an M member, 8 machine-tier twins live in those groups, 5 already agree,
    # and all 3 that disagree are the machine being wrong — `b'xgan` *bhgan for the
    # attested brhgan 把…鎖, `mq'qan` *mqekan for mkeekan 打架 (41× in speech), and
    # mg'li. Only M overrides: an attested tier (id/A/B/S/T) is evidence about the
    # exact token in hand and outranks a twin, and X is excluded for the same reason
    # as above — it has to declare itself on screen.
    V_MACHINE = set(("R", "D", "P", "E", "G", "B-rules", "C-review", "V"))
    by_bare = collections.defaultdict(dict)
    for k, rec in result.items():
        by_bare[ELISION_RE.sub("", k)][k] = rec
    for bare, twins in sorted(by_bare.items()):
        if len(twins) < 2 or any(r["tier"] == "X" for r in twins.values()):
            continue
        hand = set(r["modern"] for r in twins.values() if r["tier"] == "M")
        if len(hand) != 1:
            continue
        want = hand.pop()
        src = sorted(k for k, r in twins.items() if r["tier"] == "M")[0]
        for t in sorted(twins):
            if twins[t]["tier"] not in V_MACHINE or twins[t]["modern"] == want:
                continue
            tiers[twins[t]["tier"]] -= 1
            tiers["V"] += 1
            v_log.append((t, want, src, "M over " + twins[t]["tier"] +
                          " " + twins[t]["modern"], tokens.get(t, 0)))
            result[t] = {"modern": want, "tier": "V", "v_twin": src,
                         "v_was": twins[t]["modern"]}

    # 7c. Tier W: the written schwa before a word-initial labial. Every tier above
    # answers one token at a time, so none of them can see that a whole INITIAL
    # CLUSTER is one the orthography never writes. `seqsweep.py` asks that question
    # mechanically — a character n-gram absent from all 38,687 modern types — and the
    # largest class it returns is `^mp`, 101 map values printing a prefix spelling
    # that occurs 8 times in 277k tokens of speech. Modern writes the future/agentive
    # as `emp-` (1,651 types / 2,251 tokens) and the stative m- on a b-root as `emb-`
    # (201 / 1,011); his transcription drops the schwa, exactly as it does word-
    # internally (`xnglyeq` → hnegliq, batch 25).
    #
    # The scope is measured, not assumed, and it stops at the labials. A GENERAL
    # "value 0×, e+value attested" rule fires 144 times and is unsafe: it would take
    # `glani` → *eglani (a batch-24 word), `duk` → *eduk 門扇, `lixan` → *erihan —
    # singleton initials where the e-form is a coincidental different word. And no
    # other m-initial has the vein at all: `emn` 0 types against `mn` 1,215, `ems` 0
    # against 1,096, `emk` 5 against 920, `emg` 1 against 667, `emt` 15 against 520.
    # It is written before p and b because that is where the cluster needs breaking.
    #
    # Guarded per token, so the class evidence never overrides evidence about the
    # word in hand: his form must be unattested (both corpora) and the e-form must be
    # attested. That leaves `mpgeeguy` 偷竊者, `mpplaq` and `mputuh` 斷掉 alone, and
    # leaves `mblaiq` alone too — 5× itself, though `emblaiq` is 43×. Deliberately
    # conservative: attested is attested, the same principle as the id tier. Not a
    # manual_map entry, because freezing 128 stems there would silently override any
    # future stem fix — the map is regenerated, and a post-pass composes with it.
    # The blind half rests on class evidence alone, and the class is close to absolute:
    # exactly ONE m+labial-initial type in all 38,687 modern types lacks an e-form
    # (`mpotoh` 2×, unglossed, beside the attested `mputuh` 斷掉), and `mpa-` has 0
    # types against `empa-`'s 50 — a real prefix, "will become X" (`empatas` 100×
    # 在…讀書). Of 149 blind values, **none** takes the competing m-on-a-p-root reading:
    # not one leaves an attested p-initial word when only the m is stripped, while 100
    # leave a directly attested stem (`emp`+`nanak` 600×, `iya` 572×, `piya` 515×,
    # `baki` 488×). This is the batch-25 argument — a blind slot moves from a spelling
    # that is definitely not modern to one that is at least well formed.
    #
    # What still has to be checked is the shape of the RESULT, by the same n-gram logic
    # that found the vein: the e-form's initial must be one modern words actually take.
    # 48 of the 49 witnessless values pass with 17–455 types sharing their initial;
    # `mpyah` alone fails, `empy-` being 0×, so it stays out — his P'IYAX 來 is
    # `empiyah`, and prepending an e to a syncopated stem would only invent a new
    # impossible initial in place of the old one.
    LICIT = collections.Counter()
    for w in list(attested) + list(spoken):
        LICIT[w[:4]] += 1
    w_log = []
    for t in sorted(result):
        rec = result[t]
        v = rec["modern"]
        # X declares itself on screen; N and J are frozen populations, and this is a
        # rule about a TRUKU prefix — it has nothing to say about a man's name (`mpa`)
        # or a Japanese loan (`mbosi`), and the class branch reached all four.
        if rec["tier"] in ("X", "W", "N", "J") or v != v.lower():
            continue
        if not (v.startswith("mp") or v.startswith("mb")):
            continue
        nv = norm(v)
        if nv in attested or spoken.get(nv):
            continue                        # his form is itself modern Truku
        e = "e" + v
        ne = norm(e)
        if ne in attested or spoken.get(ne):
            how = "twin"
        elif LICIT[ne[:4]]:
            how = "class"                   # no twin, but a licit modern initial
        else:
            continue
        tiers[rec["tier"]] -= 1
        tiers["W"] += 1
        w_log.append((t, e, v, rec["tier"] + "/" + how,
                      spoken.get(ne, 0), tokens.get(t, 0)))
        result[t] = {"modern": e, "tier": "W", "w_was": v, "w_how": how}

    unmapped.sort(key=lambda x: -x[1])
    json.dump({"map": result, "review": review, "unmapped_top": unmapped[:400]},
              open(os.path.join(HERE, "modern_map.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    lines = []
    for t, rec in sorted(result.items()):
        lines.append('"%s":"%s"' % (t, rec["modern"]))
    with open(os.path.join(BASE, "site", "modern_map.js"), "w", encoding="utf-8", newline="\n") as f:
        f.write(
            "// Generated by tools/orthography/build_modern_map.py — do not edit by hand.\n"
            "// token (lowercase, Pecoraro spelling) -> modern Truku spelling, for the\n"
            "// display-only modern-spelling toggle. Tiers: identity-attested, gloss-\n"
            "// confirmed (A), unique-candidate (B), root-projected (P). Ambiguous cases stay unmapped and\n"
            "// fall through to the character rules. Regenerate after entries.js changes.\n"
            "window.MODERN_MAP = {\n" + ",\n".join(lines) + "\n};\n"
            "// Tier X: the modern entry is a DIFFERENT WORD, not a respelling. The app\n"
            "// prints these as MODERN (his original) so the toggle never quietly puts a\n"
            "// word in his mouth. Source: tools/orthography/lexical_map.json.\n"
            "window.LEXICAL_SUBS = {\n"
            + ",\n".join('"%s":1' % t for t in sorted(lexical)) + "\n};\n"
        )

    print("tokens considered:", len(tokens))
    print("tier counts:", dict(tiers))
    print("mapped:", len(result), " review:", len(review))
    print("projection: %d mapped (of which %d attested), %d ambiguous-skipped"
          % (tiers["P"], proj_att, proj_ambig))
    print("keep-l guard: %d tokens frozen to keep-l (would have been wrongly l>r'd)" % kl)
    print("relative inheritance: %d mapped, %d ambiguous-skipped, %d gloss-vetoed"
          % (tiers["R"], r_ambig, r_veto))
    print("morphology (D): %d mapped — %s" % (tiers["D"], dict(d_rules)))
    with open(os.path.join(HERE, "tier_d_log.txt"), "w", encoding="utf-8", newline="\n") as f:
        f.write("# tier D: token -> modern, rule, base, tier the base came from\n")
        for t, mod, rule, base, bt in sorted(d_log):
            f.write("%-16s %-16s %-6s %-16s %s\n" % (t, mod, rule, base, bt))
    print("example projection (E): %d mapped" % tiers["E"])
    with open(os.path.join(HERE, "tier_e_log.txt"), "w", encoding="utf-8", newline="\n") as f:
        f.write("# tier E: sentence-only token -> modern, the stem it inherited, that stem's modern form\n")
        for t, mod, sp, sm, how in sorted(e_log):
            f.write("%-16s %-16s %-14s %-14s %s\n" % (t, mod, sp, sm, how))
    print("spoken-corpus attestation (S): %d mapped" % tiers["S"])
    with open(os.path.join(HERE, "tier_s_log.txt"), "w", encoding="utf-8", newline="\n") as f:
        f.write("# tier S: his token -> the one rule-consistent reading found in transcribed Truku speech\n")
        f.write("# columns: token, modern, times in the 277k-token spoken corpus, times on screen\n")
        for t, m, c, o in sorted(s_log, key=lambda r: -r[3]):
            f.write("%-16s %-16s %6d %5d\n" % (t, m, c, o))
    print("proper names (N): %d frozen against l>r" % tiers["N"])
    with open(os.path.join(HERE, "tier_n_log.txt"), "w", encoding="utf-8", newline="\n") as f:
        f.write("# tier N: capitalized mid-sentence, never lowercase anywhere -> a name, so l stays l\n")
        for t, m, mc, o in sorted(n_log, key=lambda r: -r[3]):
            f.write("%-16s %-16s midcap=%-4d occ=%d\n" % (t, m, mc, o))
    j_rows = [(t, r["modern"], r["j_how"], tokens[t]) for t, r in result.items()
              if r["tier"] == "J"]
    print("japanese/chinese loans (J): %d romanized (%d kept on gloss evidence)"
          % (len(j_rows), sum(1 for r in j_rows if r[2] == "gloss")))
    with open(os.path.join(HERE, "tier_j_log.txt"), "w", encoding="utf-8", newline="\n") as f:
        f.write("# tier J: words tagged [emprunt jap./chin.], romanized as a class\n")
        f.write("# how = gloss: an attested modern word whose Chinese gloss agrees\n")
        f.write("#       rule : his spelling with o>u, x>h, final -e>-i, -ai>-ay,\n"
                "#              -wi>-uy; l left alone\n")
        f.write("#       base : the rule on the prefix + a base resolved above\n")
        f.write("#       hand : a curated mapping that changed something\n")
        f.write("# columns: token, modern, how, occurrences\n")
        for t, mod, how, o in sorted(j_rows, key=lambda r: (r[2], r[0])):
            f.write("%-16s %-16s %-6s %5d\n" % (t, mod, how, o))
    print("cross-entry root projection (G): %d mapped" % tiers["G"])
    with open(os.path.join(HERE, "tier_g_log.txt"), "w", encoding="utf-8", newline="\n") as f:
        f.write("# tier G: token -> modern, the ROOT stem it inherited (from any entry), that root's modern form\n")
        f.write("# columns: token, modern, stem, stem modern, occurrences, attested?\n")
        for t, mod, sp, sm, o, how in sorted(g_log, key=lambda r: -r[4]):
            f.write("%-16s %-16s %-14s %-14s %5d %s\n" % (t, mod, sp, sm, o, how))
    print("elision-mark variants (V): %d mapped" % tiers["V"])
    with open(os.path.join(HERE, "tier_v_log.txt"), "w", encoding="utf-8", newline="\n") as f:
        f.write("# tier V: the same word with his elision mark in a different place\n")
        f.write("# columns: token, modern, the twin it inherited from, that twin's tier, occurrences\n")
        for t, mod, src, st, o in sorted(v_log, key=lambda r: -r[4]):
            f.write("%-16s %-16s %-16s %-4s %5d\n" % (t, mod, src, st, o))
    print("word-initial schwa before a labial (W): %d mapped" % tiers["W"])
    with open(os.path.join(HERE, "tier_w_log.txt"), "w", encoding="utf-8", newline="\n") as f:
        f.write("# tier W: modern writes the schwa before a word-initial labial (emp-/emb-)\n")
        f.write("# columns: token, modern, the value it replaced, that value's tier, e-form spoken freq, occurrences\n")
        for t, mod, was, st, ne, o in sorted(w_log, key=lambda r: (-r[4], r[0])):
            f.write("%-16s %-17s %-16s %-8s %5d %5d\n" % (t, mod, was, st, ne, o))
    changed = sum(1 for t, r in result.items() if r["modern"] != t)
    print("mapped with actual spelling change:", changed)

if __name__ == "__main__":
    main()
